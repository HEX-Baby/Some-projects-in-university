import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
file_path = r'C:\Users\jay\Desktop\概率论与数理统计2025春.xlsx'
output_path = r'C:\Users\jay\Desktop\概率论与数理统计2025春-0992.xlsx'

df = pd.read_excel(file_path)

full_score = [100, 100, 100, 100, 50, 100, 100, 100]

all_full_score = sum(full_score)

df['平时分'] = 0

for i, c in enumerate(full_score):
    df['平时分'] += df[i + 1] / c * 100

df['平时分'] /= 8

df['平时分'] = df['平时分'].round().astype(int)

df.to_excel(output_path, index=False)

# 使用openpyxl移除第一行的加粗和边框
wb = load_workbook(output_path)
ws = wb.active

# 创建无边框样式和普通字体
no_border = Border(left=Side(style='none'),
                   right=Side(style='none'),
                   top=Side(style='none'),
                   bottom=Side(style='none'))
normal_font = Font(bold=False)

# 修改第一行的样式
for cell in ws[1]:
    cell.font = normal_font
    cell.border = no_border

wb.save(output_path)